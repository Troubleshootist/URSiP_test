from pprint import pprint

from openpyxl import load_workbook
from openpyxl.cell import Cell
from models import *


class XlsxParser:
    data_rows = []
    data_types = {}
    periods = {}

    def __init__(self, file_path):
        self.file = file_path

    def parse_xlsx(self):
        wb = load_workbook(self.file)
        ws = wb.active
        self.__parse_data_types(ws)
        self.__parse_periods(ws)
        for row in ws.iter_rows(min_row=4):
            company = Company(row[1].value)
            for cell in row:
                if 2 < cell.col_idx < 11 and cell.col_idx % 2 != 0:
                    self.__append_data_row(period=self.periods[cell.col_idx],
                                           company=company,
                                           data_type=self.data_types[cell.col_idx],
                                           cell=cell,
                                           row=row)

    def __parse_data_types(self, ws):
        for cell in ws[2]:
            if cell.value:
                self.data_types[cell.col_idx] = DataType(cell.value)
                self.data_types[cell.col_idx + 1] = DataType(cell.value)

    def __parse_periods(self, ws):
        for cell in ws[1]:
            if cell.value and cell.col_idx > 2:
                for index in range(cell.col_idx, cell.col_idx + 4):
                    self.periods[index] = Period(cell.value)

    def __append_data_row(self, period: Period, company: Company, data_type: DataType, cell: Cell, row: tuple):
        self.data_rows.append(DataRow(
            period=period,
            company=company,
            data_type=data_type,
            data={
                'data1': cell.value,
                'data2': row[cell.col_idx].value
            }
        ))
